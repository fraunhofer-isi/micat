# © 2024 - 2025 Fraunhofer-Gesellschaft e.V., München
#
# SPDX-License-Identifier: AGPL-3.0-or-later

import io
from copy import copy

import numpy as np
import openpyxl

from micat.calculation import extrapolation
from micat.calculation.ecologic import fuel_split
from micat.calculation.economic import investment, population
from micat.calculation.social import dwelling
from micat.input.data_source import DataSource
from micat.table.table import Table
from micat.utils import api


def measure_specific_parameters_template(
    request,
    database,
    confidential_database,
):
    template_args = _template_args(request)
    template_bytes = _measure_specific_template(
        template_args,
        database,
        confidential_database,
    )
    return template_bytes


# pylint: disable=unused-argument
def _measure_specific_template(
    template_args,
    database,
    confidential_database,
):
    workbook = openpyxl.load_workbook("template/measure_specific_parameters_template.xlsx")
    _fill_measure_specific_template(
        workbook,
        template_args,
        database,
        confidential_database,
    )
    output = io.BytesIO()
    workbook.save(output)

    return output


# pylint:disable=too-many-locals
def _fill_measure_specific_template(
    workbook,
    template_args,
    database,
    confidential_database,
):
    main_sheet = workbook["main"]
    affected_fuels_sheet = workbook["affectedFuels"]
    fuel_switch_sheet = workbook["fuelSwitch"]
    residential_sheet = workbook["residential"]
    context_sheet = workbook["context"]

    measure = template_args["measure"]
    population_of_municipality = population.population_of_municipality(measure)

    id_region = int(template_args["id_region"])
    context = {
        "id_mode": int(template_args["id_mode"]),
        "id_region": id_region,
        "id_subsector": int(measure["subsector"]["id"]),
        "id_action_type": int(measure["action_type"]["id"]),
        "unit": measure["unit"],
        "population_of_municipality": population_of_municipality,
    }

    final_energy_saving_by_action_type = _final_energy_saving_by_action_type(measure, context)

    data_source = DataSource(
        database, id_region, confidential_database, global_parameters=template_args["global_parameters"]
    )
    df = data_source.mapping_table("mapping__subsector__sector")._data_frame
    id_sector = df.loc[df["id_subsector"] == int(template_args["id_subsector"])]["id_sector"].item()

    wuppertal_parameters = _wuppertal_parameters(
        context,
        final_energy_saving_by_action_type,
        data_source,
    )

    _fill_main_and_fuel_sheet(
        main_sheet,
        affected_fuels_sheet,
        context,
        final_energy_saving_by_action_type,
        wuppertal_parameters,
        data_source,
    )

    years = final_energy_saving_by_action_type.years
    _fill_fuel_switch_sheet(fuel_switch_sheet, years, id_sector)

    _fill_residential_sheet(
        residential_sheet,
        context,
        final_energy_saving_by_action_type,
        wuppertal_parameters,
        data_source,
    )

    _fill_context_sheet(context_sheet, context)


def _wuppertal_parameters(
    context,
    final_energy_saving_by_action_type,
    data_source,
):
    id_region = context["id_region"]
    wuppertal_parameters_raw = data_source.table("wuppertal_parameters", {"id_region": str(id_region)})

    years = final_energy_saving_by_action_type.years
    wuppertal_parameters = extrapolation.extrapolate(wuppertal_parameters_raw, years)
    return wuppertal_parameters


def _fill_main_and_fuel_sheet(
    main_sheet,
    affected_fuels_sheet,
    context,
    final_energy_saving_by_action_type,
    wuppertal_parameters,
    data_source,
):
    years = final_energy_saving_by_action_type.years
    main_sheet = _interpolate_annual_data(main_sheet, years)
    main_sheet = _fill_unit(main_sheet, context)
    affected_fuels_sheet = _interpolate_annual_data(affected_fuels_sheet, years)
    affected_fuels_sheet = _fill_unit(affected_fuels_sheet, context)

    _fill_annual_savings(main_sheet, final_energy_saving_by_action_type)

    _fill_investment_cost(
        main_sheet,
        final_energy_saving_by_action_type,
        data_source,
    )

    _fill_subsidy_rate(main_sheet, wuppertal_parameters)

    _fill_lifetime(main_sheet, context, data_source)

    _fill_share_affected(
        affected_fuels_sheet,
        context,
        final_energy_saving_by_action_type,
        data_source,
    )


def _fill_fuel_switch_sheet(sheet, years, id_sector):
    # Remove rows, which doesn't match the sector
    rows_to_delete = []
    for idx, row in enumerate(sheet.iter_rows(min_row=2)):
        if row[1].value != id_sector:
            rows_to_delete.append(idx)
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx + 2, 1)
    # Interpolate the annual data
    _interpolate_annual_data(sheet, years)


def _fill_residential_sheet(
    sheet,
    context,
    final_energy_saving_by_action_type,
    wuppertal_parameters,
    data_source,
):
    years = final_energy_saving_by_action_type.years

    sheet = _interpolate_annual_data(sheet, years)

    _fill_number_of_affected_dwellings(
        sheet,
        context,
        final_energy_saving_by_action_type,
        data_source,
    )

    _fill_energy_poverty_targeteness(sheet, wuppertal_parameters)

    _fill_dwelling_stock(
        sheet,
        context,
        final_energy_saving_by_action_type,
        data_source,
    )

    # _fill_annual_renovation_rate(sheet, ...)   # not yet available from sqlite databases
    _fill_average_hh_per_building(sheet, wuppertal_parameters)
    _fill_average_rent(sheet, wuppertal_parameters)
    _fill_rent_premium(sheet, wuppertal_parameters)


def _fill_context_sheet(
    sheet,
    context,
):
    id_region = context["id_region"]
    sheet.cell(column=2, row=1, value=id_region)

    id_subsector = context["id_subsector"]
    sheet.cell(column=2, row=2, value=id_subsector)

    id_action_type = context["id_action_type"]
    sheet.cell(column=2, row=3, value=id_action_type)

    population_of_municipality = context["population_of_municipality"]
    sheet.cell(column=2, row=4, value=population_of_municipality)


def _annual_columns(sheet):
    annual_columns = []
    for column in sheet.columns:
        header_cell = column[0]
        value = header_cell.value
        if isinstance(value, int):
            annual_columns.append(column)
    return annual_columns


def _cell_style(cell):
    # noinspection PyProtectedMember
    style = copy(cell._style)  # pylint: disable = protected-access
    return style


def _cell_value(cell):
    value = cell.value
    if value == "=NA()":
        return float("NaN")
    else:
        return value


def _columns_to_table(columns):
    data = {}
    for column in columns:
        column_values = [_cell_value(cell) for cell in column]
        header = column_values.pop(0)
        data[header] = column_values
    table = Table(data)
    return table


def _extrapolated_nan_table(raw_annual_table, years):
    number_of_rows = len(raw_annual_table)
    data = {}
    for year in years:
        values = np.empty(number_of_rows)
        values.fill(np.nan)
        data[str(year)] = values
    annual_table = Table(data)
    return annual_table


def _fill_annual_data(
    sheet,
    annual_table,
    cell_styles,
):
    column_offset = sheet.max_column + 1

    _, column_names, _ = annual_table.column_names
    column_index = column_offset
    header_style = cell_styles[0]
    for column_name in column_names:
        cell = sheet.cell(row=1, column=column_index)
        cell.value = column_name
        cell._style = header_style  # pylint: disable=protected-access
        column_index += 1

    for row_index, row in annual_table.iterrows():
        column_index = column_offset
        cell_style = cell_styles[row_index + 1]
        for value in row:
            cell = sheet.cell(row=row_index + 2, column=column_index)
            cell.value = round(value, 2)
            cell._style = cell_style  # pylint: disable=protected-access
            column_index += 1

    return sheet


def _fill_annual_savings(sheet, final_energy_saving_by_action_type):
    _fill_annual_series(sheet, 7, 2, final_energy_saving_by_action_type)


def _fill_annual_series(
    sheet,
    start_column_index,
    row_index,
    series_or_table,
):
    if isinstance(series_or_table, Table):
        values = series_or_table.values[0]
    else:
        values = series_or_table.values
    column_index = start_column_index
    for value in values:
        cell = sheet.cell(column=column_index, row=row_index)
        cell.value = round(value, 2)
        column_index += 1


def _fill_table(
    sheet,
    start_column_index,
    start_row_index,
    table,
):
    row_index = start_row_index
    for _index, row in table.iterrows():
        column_index = start_column_index
        for value in row.values:
            cell = sheet.cell(column=column_index, row=row_index)
            cell.value = round(value, 2)
            column_index += 1
        row_index += 1


def _fill_dwelling_stock(
    sheet,
    context,
    final_energy_saving_by_action_type,
    data_source,
):
    id_region = context["id_region"]
    population_of_municipality = context["population_of_municipality"]

    dwelling_stock = dwelling.dwelling_stock(
        final_energy_saving_by_action_type,
        data_source,
        id_region,
        population_of_municipality,
    )

    _fill_annual_series(sheet, 5, 4, dwelling_stock)


def _fill_average_hh_per_building(sheet, wuppertal_parameters):
    average_hh_per_building = wuppertal_parameters.reduce("id_parameter", 31)
    _fill_annual_series(sheet, 5, 6, average_hh_per_building)


def _fill_average_rent(sheet, wuppertal_parameters):
    average_rent = wuppertal_parameters.reduce("id_parameter", 29)
    _fill_annual_series(sheet, 5, 7, average_rent)


def _fill_rent_premium(sheet, wuppertal_parameters):
    rent_premium = wuppertal_parameters.reduce("id_parameter", 34)
    _fill_annual_series(sheet, 5, 8, rent_premium)


def _fill_energy_poverty_targeteness(sheet, wuppertal_parameters):
    energy_poverty_targeteness = wuppertal_parameters.reduce("id_parameter", 25)
    _fill_annual_series(sheet, 5, 3, energy_poverty_targeteness)


def _fill_investment_cost(
    sheet,
    final_energy_saving_by_action_type,
    data_source,
):
    investment_cost = investment.investment_cost_in_euro(
        final_energy_saving_by_action_type,
        data_source,
    )
    # Convert to million euros to display the advanced parameter investment cost in mio. €
    investment_cost._data_frame = investment_cost._data_frame.select_dtypes(exclude=["object", "datetime"]) / 1_000_000
    _fill_annual_series(sheet, 7, 3, investment_cost)


def _fill_lifetime(sheet, context, database):
    lifetime = _lifetime(context, database)
    cell = sheet.cell(column=6, row=5)
    cell.value = lifetime


def _lifetime(context, database):
    id_subsector = context["id_subsector"]
    id_action_type = context["id_action_type"]
    wuppertal_sector_parameters = database.table(
        "wuppertal_sector_parameters", {"id_subsector": str(id_subsector), "id_action_type": str(id_action_type)}
    )
    lifetime = wuppertal_sector_parameters.reduce("id_parameter", 36)
    return lifetime


def _fill_share_affected(
    sheet,
    context,
    final_energy_saving_by_action_type,
    data_source,
):
    id_mode = context["id_mode"]
    id_region = context["id_region"]
    id_subsector = context["id_subsector"]
    subsector_ids = [id_subsector]

    share_affected = fuel_split.fuel_split_by_action_type(
        final_energy_saving_by_action_type,
        data_source,
        id_mode,
        id_region,
        subsector_ids,
    )

    _fill_table(sheet, 7, 2, share_affected)


def _fill_subsidy_rate(sheet, wuppertal_parameters):
    subsidy_rate = wuppertal_parameters.reduce("id_parameter", 35)
    _fill_annual_series(sheet, 7, 4, subsidy_rate)


# pylint: disable=duplicate-code
def _fill_number_of_affected_dwellings(
    sheet,
    context,
    final_energy_saving_by_action_type,
    data_source,
):
    id_region = context["id_region"]
    population_of_municipality = context["population_of_municipality"]

    number_of_affected_dwellings = dwelling.number_of_affected_dwellings(
        final_energy_saving_by_action_type,
        data_source,
        id_region,
        population_of_municipality,
    )
    _fill_annual_series(sheet, 5, 2, number_of_affected_dwellings)


def _fill_unit(sheet, context):
    unit_object = context["unit"]
    unit = unit_object["symbol"]
    sheet["D2"] = unit
    return sheet


def _interpolate_annual_data(sheet, years):
    sheet_without_annual_data, raw_annual_table, cell_styles = _split_sheet(sheet)
    if raw_annual_table.contains_non_nan():
        annual_table = extrapolation.extrapolate(raw_annual_table, years)
    else:
        annual_table = _extrapolated_nan_table(raw_annual_table, years)
    sheet = _fill_annual_data(sheet_without_annual_data, annual_table, cell_styles)

    return sheet


def _final_energy_saving_by_action_type(measure, context):
    annual_data = {key: value for key, value in measure.items() if key.isdigit()}
    annual_data["id_measure"] = measure["id"]
    annual_data["id_subsector"] = context["id_subsector"]
    annual_data["id_action_type"] = context["id_action_type"]

    savings = Table([annual_data])
    return savings


def _split_sheet(sheet):
    annual_columns = _annual_columns(sheet)
    annual_table = _columns_to_table(annual_columns)

    first_column = annual_columns[0]
    cell_styles = list(map(_cell_style, first_column))

    first_column_index = annual_columns[0][0].column
    number_of_columns = len(annual_columns)
    sheet.delete_cols(first_column_index, number_of_columns)

    return sheet, annual_table, cell_styles


def _template_args(request):
    query_dict = api.parse_request(request)
    measure = query_dict["json"]
    args = {
        "measure": measure,
        "id_mode": query_dict["id_mode"],
        "id_region": query_dict["id_region"],
        "id_subsector": query_dict["id_subsector"],
        "parameter_table_names": [
            "eurostat_final_parameters",
            "eurostat_final_sector_parameters",
            "eurostat_primary_parameters",
            "iiasa_final_subsector_parameters",
            "mixed_final_constant_parameters",
            "primes_parameters",
            "primes_primary_parameters",
        ],
        "global_parameters": measure.get("global_parameters"),
    }
    return args
